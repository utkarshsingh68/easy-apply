"""
main.py
-------
FastAPI application — the single entry point for the AI Job Application Agent.

Available endpoints
───────────────────
POST /add-job                  Scrape a job URL and store company data.
POST /generate-email/{id}      Generate a personalised email via LLM.
POST /send-email/{id}          Send the email via Gmail SMTP.
GET  /emails                   List all email logs.
GET  /companies                List all stored companies.

Run with:
    uvicorn main:app --host 0.0.0.0 --port 8000 --reload
"""

from __future__ import annotations

import csv
import json
import random
import re
import time
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path as FilePath
from urllib.parse import parse_qs, urlparse
from typing import Any

import httpx
from fastapi import FastAPI, File, Form, HTTPException, Path, Query, Request, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel, field_validator

import supabase_client as db
from ai_email_generator import generate_email
from config import settings
from email_sender import send_email
from scraper import scrape_job
from utils import is_valid_email, logger


BASE_DIR = FilePath(__file__).resolve().parent
FRONTEND_DIST_DIR = BASE_DIR / "frontend" / "dist"
FRONTEND_ASSETS_DIR = FRONTEND_DIST_DIR / "assets"

TEMPLATE_LIBRARY: dict[str, str] = {
    "default": "Balanced and concise outreach style.",
    "value-first": "Lead with measurable outcomes and strongest skill-match.",
    "story": "Use a short narrative from motivation to contribution.",
}
FOLLOWUP_QUEUE: list[dict[str, Any]] = []
REPLY_EVENTS: list[dict[str, Any]] = []


# ── Lifespan (startup / shutdown) ─────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Verify Supabase connectivity on startup."""
    try:
        db.list_companies()
        logger.info("[Startup] Supabase connection OK.")
    except Exception as exc:
        logger.error(f"[Startup] Supabase connection failed: {exc}")
        # Don't crash — operator may fix env vars and reload without restart
    yield
    logger.info("[Shutdown] AI Job Agent stopped.")


# ── App factory ───────────────────────────────────────────────────────────────

app = FastAPI(
    title="AI Job Application Agent",
    description=(
        "Automatically scrapes job postings, generates personalised outreach "
        "emails with an LLM, and sends them via Gmail SMTP. "
        "All data is persisted in Supabase."
    ),
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # Tighten in production
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Request / Response schemas ────────────────────────────────────────────────

class AddJobRequest(BaseModel):
    """Payload for POST /add-job (URL scraping OR manual company entry)."""

    job_url: str | None = None
    company_name: str | None = None
    job_title: str = "Unknown Position"
    email: str = ""

    @field_validator("job_url", mode="before")
    @classmethod
    def normalise_url(cls, v: Any) -> Any:
        if v is None:
            return None
        url = str(v).strip()
        if not url:
            return None
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        return url

    @field_validator("company_name", mode="before")
    @classmethod
    def normalise_company_name(cls, v: Any) -> Any:
        if v is None:
            return None
        text = str(v).strip()
        return text or None


class CompanyResponse(BaseModel):
    id: str
    company_name: str
    job_title: str
    email: str
    job_url: str
    created_at: str | None = None


class EmailLogResponse(BaseModel):
    id: str
    company_id: str
    email_subject: str
    email_body: str
    status: str
    sent_at: str | None = None


class GenerateEmailResponse(BaseModel):
    log_id: str
    company_id: str
    subject: str
    body: str
    status: str


class SendEmailResponse(BaseModel):
    log_id: str
    company_id: str
    status: str
    message: str


class HrEmailRequest(BaseModel):
    hr_email: str | None = None
    template_name: str = "default"
    jd_text: str = ""
    resume_profile: str | None = None

    @field_validator("hr_email", mode="before")
    @classmethod
    def normalise_hr_email(cls, v: Any) -> Any:
        if v is None:
            return None
        text = str(v).strip().lower()
        return text or None

    @field_validator("template_name", mode="before")
    @classmethod
    def normalise_template_name(cls, v: Any) -> str:
        if v is None:
            return "default"
        text = str(v).strip().lower()
        return text or "default"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_company_or_404(company_id: str) -> dict[str, Any]:
    """Fetch a company row or raise HTTP 404."""
    company = db.get_company_by_id(company_id)
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Company with id={company_id!r} not found.",
        )
    return company


def _synthetic_job_url(company_name: str) -> str:
    """Generate a unique pseudo-URL for manual entries."""
    slug = re.sub(r"[^a-z0-9]+", "-", company_name.lower()).strip("-") or "company"
    return f"manual://{slug}-{int(time.time() * 1000)}"


def _has_existing_email_for_company_name(company_name: str) -> bool:
    """Return True if this company already has a pending/sent email log."""
    for row in db.list_email_logs():
        company = row.get("companies") or {}
        existing_name = str(company.get("company_name", "")).strip().lower()
        if existing_name == company_name.strip().lower() and row.get("status") in {"pending", "sent"}:
            return True
    return False


def _resolve_hr_email(company: dict[str, Any], payload: HrEmailRequest | None) -> str:
    """Get HR email from request payload first, then stored company data."""
    hr_email = (payload.hr_email if payload else None) or company.get("email", "")
    return str(hr_email).strip().lower()


def _require_send_authorization(request: Request) -> None:
    """Guard live send endpoints from public abuse.

    If `public_send_enabled` is False, callers must provide `X-Admin-Token`
    matching `send_admin_token` in `.env`.
    """
    if settings.public_send_enabled:
        return

    configured_token = str(settings.send_admin_token or "").strip()
    provided_token = str(request.headers.get("x-admin-token", "")).strip()

    if not configured_token:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "Sending is locked. Set SEND_ADMIN_TOKEN in .env and provide it "
                "in X-Admin-Token header."
            ),
        )
    if provided_token != configured_token:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Unauthorized send action. Invalid or missing owner token.",
        )


def _normalise_job_url(job_url: str) -> str:
    """Ensure job URLs are absolute HTTP(S) URLs."""
    url = str(job_url).strip()
    if url and not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def _parse_bulk_upload_rows(filename: str, raw_bytes: bytes) -> list[tuple[int, dict[str, str]]]:
    """Parse CSV/XLSX rows into normalized dictionaries with source row numbers."""
    normalized_name = filename.lower()
    rows_with_numbers: list[tuple[int, dict[str, str]]] = []

    if normalized_name.endswith(".csv"):
        try:
            csv_text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV must be UTF-8 encoded.",
            )

        reader = csv.DictReader(StringIO(csv_text))
        if not reader.fieldnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV header row is missing.",
            )

        for row_number, row in enumerate(reader, start=2):
            normalized_row = {
                str(k).strip().lower(): str(v).strip() if v is not None else ""
                for k, v in row.items()
                if k is not None
            }
            if any(value for value in normalized_row.values()):
                rows_with_numbers.append((row_number, normalized_row))
        return rows_with_numbers

    if normalized_name.endswith(".xlsx") or normalized_name.endswith(".xlsm"):
        try:
            workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True, read_only=True)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file: {exc}",
            )

        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            workbook.close()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel header row is missing.",
            )

        normalized_headers = [str(h).strip().lower() if h is not None else "" for h in headers]
        if not any(normalized_headers):
            workbook.close()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel header row is empty.",
            )

        for row_number, row_values in enumerate(iterator, start=2):
            normalized_row: dict[str, str] = {}
            for index, cell_value in enumerate(row_values):
                if index >= len(normalized_headers):
                    break
                key = normalized_headers[index]
                if not key:
                    continue
                normalized_row[key] = str(cell_value).strip() if cell_value is not None else ""

            if any(value for value in normalized_row.values()):
                rows_with_numbers.append((row_number, normalized_row))

        workbook.close()
        return rows_with_numbers

    if normalized_name.endswith(".xls"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Legacy .xls is not supported. Save as .xlsx or .csv and upload again.",
        )

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=(
            "Unsupported file type. Upload .csv, .xlsx, or .xlsm "
            "with columns: job_url, company_name, job_title, hr_email."
        ),
    )


def _apply_mapping(row: dict[str, str], mapping: dict[str, str] | None) -> dict[str, str]:
    """Map arbitrary source column names to canonical names."""
    if not mapping:
        return row

    mapped = dict(row)
    for canonical_key, source_key in mapping.items():
        source = str(source_key or "").strip().lower()
        if not source:
            continue
        mapped[canonical_key] = row.get(source, "")
    return mapped


def _parse_mapping_json(mapping_json: str | None) -> dict[str, str] | None:
    if not mapping_json:
        return None
    try:
        mapping_raw = json.loads(mapping_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid mapping_json: {exc}",
        )

    if not isinstance(mapping_raw, dict):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="mapping_json must be a JSON object.",
        )

    return {
        str(k).strip().lower(): str(v).strip().lower()
        for k, v in mapping_raw.items()
        if str(k).strip()
    }


def _google_sheet_to_csv_url(sheet_url: str) -> str:
    """Convert a Google Sheets URL to direct CSV export URL."""
    parsed = urlparse(sheet_url)
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", parsed.path)
    if not match:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Google Sheets URL.",
        )

    sheet_id = match.group(1)
    query = parse_qs(parsed.query)
    gid = query.get("gid", ["0"])[0]
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


async def _parse_rows_from_input(file: UploadFile | None, google_sheet_url: str | None) -> list[tuple[int, dict[str, str]]]:
    """Load rows from uploaded file or Google Sheets URL."""
    if file is not None:
        filename = file.filename or ""
        raw_bytes = await file.read()
        if not raw_bytes:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )
        return _parse_bulk_upload_rows(filename, raw_bytes)

    if google_sheet_url:
        csv_url = _google_sheet_to_csv_url(google_sheet_url)
        async with httpx.AsyncClient(timeout=25.0) as client:
            response = await client.get(csv_url)
            response.raise_for_status()
            return _parse_bulk_upload_rows("sheet.csv", response.content)

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Provide either file or google_sheet_url.",
    )


def _extract_jd_signals(jd_text: str, job_title: str) -> dict[str, Any]:
    """Extract lightweight personalization signals from JD text."""
    text = f"{job_title} {jd_text}".lower()
    skill_candidates = ["python", "fastapi", "react", "sql", "aws", "docker", "kubernetes", "ml", "llm"]
    responsibility_candidates = ["build", "design", "optimize", "deploy", "mentor", "collaborate"]
    matched_skills = [skill for skill in skill_candidates if re.search(rf"\b{re.escape(skill)}\b", text)]
    responsibilities = [w for w in responsibility_candidates if re.search(rf"\b{re.escape(w)}\w*\b", text)]

    seniority = "mid"
    if any(x in text for x in ["senior", "lead", "staff", "principal"]):
        seniority = "senior"
    elif any(x in text for x in ["intern", "junior", "entry"]):
        seniority = "junior"

    location_match = re.search(r"\b(remote|hybrid|onsite|on-site)\b", text)
    location = location_match.group(1) if location_match else "unknown"

    return {
        "matched_skills": matched_skills,
        "responsibilities": responsibilities[:4],
        "seniority": seniority,
        "location": location,
    }


def _pick_template(template_a: str, template_b: str, row_number: int) -> str:
    """Choose template for A/B testing with deterministic row split."""
    a = (template_a or "default").strip().lower()
    b = (template_b or "default").strip().lower()
    if a not in TEMPLATE_LIBRARY:
        a = "default"
    if b not in TEMPLATE_LIBRARY:
        b = "default"
    return a if row_number % 2 == 0 else b


def _guess_contact_emails(company_name: str, job_url: str) -> list[str]:
    """Generate candidate recruiter emails using domain heuristics."""
    domain = ""
    if job_url.startswith(("http://", "https://")):
        netloc = urlparse(job_url).netloc.lower()
        domain = netloc.split(":")[0]
        if domain.startswith("www."):
            domain = domain[4:]

    if not domain:
        slug = re.sub(r"[^a-z0-9]+", "", company_name.lower())
        if slug:
            domain = f"{slug}.com"

    if not domain:
        return []

    prefixes = ["hr", "careers", "recruitment", "talent", "jobs"]
    return [f"{prefix}@{domain}" for prefix in prefixes]


def _resolve_resume_path(job_title: str, jd_text: str, resume_profile: str | None) -> str | None:
    """Pick a resume attachment path based on explicit profile or role signals."""
    resume_dir = BASE_DIR / "resume"
    if resume_profile:
        candidate = (resume_dir / resume_profile).resolve()
        if candidate.exists() and candidate.is_file():
            return str(candidate)

    text = f"{job_title} {jd_text}".lower()
    profile_candidates: list[tuple[str, str]] = [
        ("data", "resume_data.pdf"),
        ("ml", "resume_ml.pdf"),
        ("backend", "resume_backend.pdf"),
        ("frontend", "resume_frontend.pdf"),
    ]
    for keyword, filename in profile_candidates:
        if keyword in text:
            candidate = (resume_dir / filename).resolve()
            if candidate.exists() and candidate.is_file():
                return str(candidate)

    return None


def _schedule_followups(
    company_id: str,
    recipient_email: str,
    email_subject: str,
    email_body: str,
    followup_days: list[int],
) -> None:
    now = datetime.now(timezone.utc)
    for index, day in enumerate(followup_days, start=1):
        FOLLOWUP_QUEUE.append(
            {
                "id": f"fu-{company_id}-{int(time.time() * 1000)}-{index}-{random.randint(100, 999)}",
                "company_id": company_id,
                "step": index,
                "recipient_email": recipient_email,
                "subject": f"Follow-up: {email_subject}",
                "body": email_body,
                "due_at": (now + timedelta(days=day)).isoformat(),
                "status": "scheduled",
            }
        )


def _parse_followup_days(raw: str) -> list[int]:
    days: list[int] = []
    for token in (raw or "").split(","):
        t = token.strip()
        if not t:
            continue
        if t.isdigit() and int(t) > 0:
            days.append(int(t))
    return days or [3, 7]


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.post(
    "/add-job",
    response_model=CompanyResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Scrape a job posting and store it in Supabase",
    tags=["Jobs"],
)
async def add_job(payload: AddJobRequest) -> CompanyResponse:
    """
    1. Scrape the provided job URL (BeautifulSoup → Playwright fallback).
    2. Extract company name, job title, and recruiter email.
    3. Store the record in the `companies` table.
    4. Return the created row.

    Returns **409 Conflict** if the job URL already exists in the database.
    """
    if not payload.job_url and not payload.company_name:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Provide either job_url or company_name.",
        )

    if payload.job_url:
        job_url = payload.job_url

        # Idempotency check before hitting the scraper
        existing = db.get_company_by_url(job_url)
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Job URL already tracked. Existing company_id={existing['id']}",
            )

        try:
            scraped = await scrape_job(job_url)
        except Exception as exc:
            logger.error(f"[API /add-job] Scraper error: {exc}")
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Scraping failed: {exc}",
            )
    else:
        company_name = payload.company_name or "Unknown Company"
        # For manual entries we use a synthetic unique identifier in job_url.
        scraped = {
            "company_name": company_name,
            "job_title": payload.job_title or "Unknown Position",
            "email": (payload.email or "").strip().lower(),
            "job_url": _synthetic_job_url(company_name),
        }

    try:
        row = db.insert_company(
            company_name=scraped["company_name"],
            job_title=scraped["job_title"],
            email=scraped["email"],
            job_url=scraped["job_url"],
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc))
    except Exception as exc:
        logger.error(f"[API /add-job] DB insert error: {exc}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {exc}",
        )

    return CompanyResponse(**row)


@app.post(
    "/bulk-upload-send",
    summary="Upload file/sheet and run bulk outreach campaign",
    tags=["Bulk"],
)
async def bulk_upload_and_send(
    request: Request,
    file: UploadFile | None = File(None),
    google_sheet_url: str | None = Form(None),
    mapping_json: str | None = Form(None),
    dry_run: bool = Form(False),
    template_a: str = Form("default"),
    template_b: str = Form("value-first"),
    followup_days: str = Form("3,7"),
    enable_enrichment: bool = Form(True),
) -> dict[str, Any]:
    """
        File headers supported:
      - job_url (optional if company_name is provided)
      - company_name (optional if job_url is provided)
      - job_title (optional)
      - hr_email (required; alias: email)
            - jd_text (optional; used for personalization)

    For each row this endpoint:
      1. Stores/fetches company data.
      2. Generates outreach drafts with JD-aware personalization.
      3. Sends emails (unless dry_run=true) and updates status.
      4. Schedules follow-ups for successful sends.
    """
    _require_send_authorization(request)

    rows = await _parse_rows_from_input(file=file, google_sheet_url=google_sheet_url)
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in uploaded file.",
        )
    mapping = _parse_mapping_json(mapping_json)
    followup_plan = _parse_followup_days(followup_days)

    results: list[dict[str, Any]] = []
    sent_count = 0
    draft_count = 0
    failed_count = 0
    skipped_count = 0

    for row_number, source_row in rows:
        normalized_row = _apply_mapping(source_row, mapping)

        job_url = _normalise_job_url(normalized_row.get("job_url", ""))
        company_name = normalized_row.get("company_name", "")
        job_title = normalized_row.get("job_title", "") or "Unknown Position"
        jd_text = normalized_row.get("jd_text", "")
        hr_email = (normalized_row.get("hr_email") or normalized_row.get("email") or "").lower()

        if not hr_email and enable_enrichment:
            guesses = _guess_contact_emails(company_name, job_url)
            hr_email = guesses[0] if guesses else ""

        if not hr_email or not is_valid_email(hr_email):
            failed_count += 1
            results.append(
                {
                    "row": row_number,
                    "status": "failed",
                    "message": "Missing or invalid hr_email (enrichment fallback exhausted).",
                }
            )
            continue

        if not job_url and not company_name:
            failed_count += 1
            results.append(
                {
                    "row": row_number,
                    "status": "failed",
                    "message": "Provide either job_url or company_name.",
                }
            )
            continue

        try:
            company: dict[str, Any] | None = None

            if job_url:
                company = db.get_company_by_url(job_url)
                if not company:
                    scraped = await scrape_job(job_url)
                    company = db.insert_company(
                        company_name=scraped["company_name"],
                        job_title=scraped["job_title"],
                        email=hr_email,
                        job_url=scraped["job_url"],
                    )
            else:
                company = db.get_company_by_name(company_name)
                if not company:
                    company = db.insert_company(
                        company_name=company_name,
                        job_title=job_title,
                        email=hr_email,
                        job_url=_synthetic_job_url(company_name),
                    )

            if company is None:
                raise RuntimeError("Unable to resolve company record.")

            if hr_email != company.get("email", ""):
                company = db.update_company_email(company["id"], hr_email)

            existing_log = db.get_email_log_by_company(company["id"])
            if existing_log and existing_log.get("status") == "sent":
                skipped_count += 1
                results.append(
                    {
                        "row": row_number,
                        "company_id": company["id"],
                        "company_name": company["company_name"],
                        "status": "skipped",
                        "message": "Email already sent for this company row.",
                    }
                )
                continue

            if not existing_log:
                if _has_existing_email_for_company_name(company["company_name"]):
                    skipped_count += 1
                    results.append(
                        {
                            "row": row_number,
                            "company_id": company["id"],
                            "company_name": company["company_name"],
                            "status": "skipped",
                            "message": "Duplicate outreach blocked for this company name.",
                        }
                    )
                    continue

                template_name = _pick_template(template_a, template_b, row_number)
                jd_signals = _extract_jd_signals(jd_text, company["job_title"])
                draft = await generate_email(
                    company_name=company["company_name"],
                    job_title=company["job_title"],
                    recipient_email=hr_email,
                    template_name=template_name,
                    jd_signals=jd_signals,
                )
                log = db.create_email_log(
                    company_id=company["id"],
                    email_subject=draft.subject,
                    email_body=draft.body,
                    status="pending",
                )
            else:
                log = existing_log

            if dry_run:
                draft_count += 1
                results.append(
                    {
                        "row": row_number,
                        "company_id": company["id"],
                        "company_name": company["company_name"],
                        "status": "draft",
                        "message": "Draft generated. dry_run=true so email was not sent.",
                    }
                )
                continue

            resume_path = _resolve_resume_path(
                job_title=company.get("job_title", ""),
                jd_text=jd_text,
                resume_profile=normalized_row.get("resume_profile") or None,
            )

            send_result = await send_email(
                to_address=hr_email,
                subject=log["email_subject"],
                body=log["email_body"],
                attach_resume=True,
                resume_path_override=resume_path,
            )
            new_status = "sent" if send_result.success else "failed"
            db.update_email_status(log["id"], new_status)

            if send_result.success:
                sent_count += 1
                _schedule_followups(
                    company_id=company["id"],
                    recipient_email=hr_email,
                    email_subject=log["email_subject"],
                    email_body=log["email_body"],
                    followup_days=followup_plan,
                )
            else:
                failed_count += 1

            results.append(
                {
                    "row": row_number,
                    "company_id": company["id"],
                    "company_name": company["company_name"],
                    "status": new_status,
                    "message": send_result.message,
                }
            )
        except Exception as exc:
            failed_count += 1
            logger.error(f"[API /bulk-upload-send] row={row_number} failed: {exc}")
            results.append(
                {
                    "row": row_number,
                    "status": "failed",
                    "message": str(exc),
                }
            )

    return {
        "total_rows": len(results),
        "sent": sent_count,
        "draft": draft_count,
        "failed": failed_count,
        "skipped": skipped_count,
        "followups_scheduled": len([f for f in FOLLOWUP_QUEUE if f.get("status") == "scheduled"]),
        "results": results,
    }


@app.post(
    "/generate-email/{company_id}",
    response_model=GenerateEmailResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Generate a personalised outreach email using an LLM",
    tags=["Email"],
)
async def generate_email_endpoint(
    company_id: str = Path(..., description="UUID of the company row"),
    payload: HrEmailRequest | None = None,
) -> GenerateEmailResponse:
    """
    1. Fetch the company record.
    2. Call the LLM (Groq or OpenAI) to draft a personalised email.
    3. Persist the draft in `email_logs` with status='pending'.
    4. Return the draft subject, body, and log ID.

    Returns **409 Conflict** if a pending/sent email log already exists for
    this company (duplicate-prevention guard).
    """
    company = _get_company_or_404(company_id)
    hr_email = _resolve_hr_email(company, payload)

    if hr_email:
        if not is_valid_email(hr_email):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Provided HR email is invalid.",
            )
        if hr_email != company.get("email", ""):
            company = db.update_company_email(company_id, hr_email)

    # Duplicate prevention by company row
    existing_log = db.get_email_log_by_company(company_id)
    if existing_log:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"An email log with status={existing_log['status']!r} already exists "
                f"for company_id={company_id}. log_id={existing_log['id']}"
            ),
        )

    # Duplicate prevention across multiple rows of the same company
    if _has_existing_email_for_company_name(company["company_name"]):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "An email is already pending/sent for this company name. "
                "Duplicate outreach is blocked."
            ),
        )

    try:
        template_name = payload.template_name if payload else "default"
        jd_text = payload.jd_text if payload else ""
        jd_signals = _extract_jd_signals(jd_text, company["job_title"])
        draft = await generate_email(
            company_name=company["company_name"],
            job_title=company["job_title"],
            recipient_email=hr_email,
            template_name=template_name,
            jd_signals=jd_signals,
        )
    except RuntimeError as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=str(exc),
        )

    try:
        log_row = db.create_email_log(
            company_id=company_id,
            email_subject=draft.subject,
            email_body=draft.body,
            status="pending",
        )
    except Exception as exc:
        logger.error(f"[API /generate-email] DB error: {exc}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {exc}",
        )

    return GenerateEmailResponse(
        log_id=log_row["id"],
        company_id=company_id,
        subject=draft.subject,
        body=draft.body,
        status="pending",
    )


@app.post(
    "/send-email/{company_id}",
    response_model=SendEmailResponse,
    summary="Send the pending outreach email via Gmail SMTP",
    tags=["Email"],
)
async def send_email_endpoint(
    request: Request,
    company_id: str = Path(..., description="UUID of the company row"),
    payload: HrEmailRequest | None = None,
) -> SendEmailResponse:
    """
    1. Fetch the company record and the most recent pending email log.
    2. Validate that a recruiter email address is available.
    3. Send via Gmail SMTP (with resume attachment if configured).
    4. Update the email_log status to 'sent' or 'failed'.

    Returns **404** if no pending email log exists (run /generate-email first).
    Returns **400** if the company has no recruiter email address.
    """
    _require_send_authorization(request)

    company = _get_company_or_404(company_id)

    # Retrieve the pending log to get subject + body
    log = db.get_email_log_by_company(company_id)
    if not log or log["status"] not in ("pending",):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "No pending email log found for this company. "
                "Call POST /generate-email/{company_id} first."
            ),
        )

    recipient = _resolve_hr_email(company, payload)
    if recipient and recipient != company.get("email", ""):
        company = db.update_company_email(company_id, recipient)

    if not recipient or not is_valid_email(recipient):
        # Update log to failed since we can't send without an address
        db.update_email_status(log["id"], "failed")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "No valid recruiter email address found for this company. "
                "Update the record manually in Supabase and retry."
            ),
        )

    jd_text = payload.jd_text if payload else ""
    resume_profile = payload.resume_profile if payload else None
    resume_path = _resolve_resume_path(company.get("job_title", ""), jd_text, resume_profile)

    result = await send_email(
        to_address=recipient,
        subject=log["email_subject"],
        body=log["email_body"],
        attach_resume=True,
        resume_path_override=resume_path,
    )

    new_status = "sent" if result.success else "failed"
    db.update_email_status(log["id"], new_status)

    if result.success:
        _schedule_followups(
            company_id=company_id,
            recipient_email=recipient,
            email_subject=log["email_subject"],
            email_body=log["email_body"],
            followup_days=[3, 7],
        )

    return SendEmailResponse(
        log_id=log["id"],
        company_id=company_id,
        status=new_status,
        message=result.message,
    )


@app.patch(
    "/companies/{company_id}/hr-email",
    response_model=CompanyResponse,
    summary="Save or update a company HR email",
    tags=["Jobs"],
)
async def update_company_hr_email(
    payload: HrEmailRequest,
    company_id: str = Path(..., description="UUID of the company row"),
) -> CompanyResponse:
    """Store a manually provided HR email for future generation and sending."""
    _get_company_or_404(company_id)
    hr_email = (payload.hr_email or "").strip().lower()
    if not hr_email or not is_valid_email(hr_email):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Provide a valid HR email address.",
        )

    row = db.update_company_email(company_id, hr_email)
    return CompanyResponse(**row)


@app.get(
    "/emails",
    response_model=list[EmailLogResponse],
    summary="List all email logs",
    tags=["Email"],
)
async def list_emails() -> list[EmailLogResponse]:
    """Return all rows from `email_logs` ordered by sent_at descending."""
    rows = db.list_email_logs()
    return [EmailLogResponse(**row) for row in rows]


@app.get(
    "/companies",
    response_model=list[CompanyResponse],
    summary="List all scraped companies",
    tags=["Jobs"],
)
async def list_companies() -> list[CompanyResponse]:
    """Return all rows from `companies` ordered by created_at descending."""
    rows = db.list_companies()
    return [CompanyResponse(**row) for row in rows]


@app.delete(
    "/companies/{company_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a company",
    tags=["Jobs"],
)
async def delete_company(company_id: str = Path(..., description="UUID of the company row")) -> Response:
    """Delete one company by id.

    Email logs linked to the company are removed by DB cascade.
    """
    _get_company_or_404(company_id)
    db.delete_company(company_id)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/templates", tags=["Campaign"], summary="List available templates")
async def list_templates() -> dict[str, Any]:
    return {
        "templates": [
            {"name": name, "guidance": guidance}
            for name, guidance in sorted(TEMPLATE_LIBRARY.items())
        ]
    }


@app.post("/templates", tags=["Campaign"], summary="Create or update a template")
async def upsert_template(
    name: str = Form(...),
    guidance: str = Form(...),
) -> dict[str, Any]:
    template_name = name.strip().lower()
    if not template_name:
        raise HTTPException(status_code=400, detail="Template name is required.")
    if not guidance.strip():
        raise HTTPException(status_code=400, detail="Template guidance is required.")
    TEMPLATE_LIBRARY[template_name] = guidance.strip()
    return {"ok": True, "name": template_name, "guidance": TEMPLATE_LIBRARY[template_name]}


@app.post("/import/preview", tags=["Bulk"], summary="Preview mapped rows before sending")
async def import_preview(
    file: UploadFile | None = File(None),
    google_sheet_url: str | None = Form(None),
    mapping_json: str | None = Form(None),
    max_rows: int = Form(20),
) -> dict[str, Any]:
    rows = await _parse_rows_from_input(file=file, google_sheet_url=google_sheet_url)
    mapping = _parse_mapping_json(mapping_json)

    preview_rows: list[dict[str, Any]] = []
    row_errors: list[dict[str, Any]] = []
    for row_number, row in rows[: max(1, min(max_rows, 100))]:
        mapped = _apply_mapping(row, mapping)
        job_url = _normalise_job_url(mapped.get("job_url", ""))
        company_name = mapped.get("company_name", "")
        hr_email = (mapped.get("hr_email") or mapped.get("email") or "").lower()

        issues: list[str] = []
        if not job_url and not company_name:
            issues.append("missing job_url/company_name")
        if not hr_email:
            issues.append("missing hr_email")
        elif not is_valid_email(hr_email):
            issues.append("invalid hr_email")

        if issues:
            row_errors.append({"row": row_number, "issues": issues})

        preview_rows.append(
            {
                "row": row_number,
                "job_url": job_url,
                "company_name": company_name,
                "job_title": mapped.get("job_title", "") or "Unknown Position",
                "hr_email": hr_email,
                "jd_text": mapped.get("jd_text", ""),
            }
        )

    detected_headers = sorted({key for _, r in rows for key in r.keys()})
    return {
        "total_rows": len(rows),
        "detected_headers": detected_headers,
        "mapping": mapping or {},
        "preview": preview_rows,
        "row_errors": row_errors,
    }


@app.get("/contact-enrich", tags=["Campaign"], summary="Suggest fallback recruiter emails")
async def contact_enrich(
    company_name: str = Query(""),
    job_url: str = Query(""),
) -> dict[str, Any]:
    guesses = _guess_contact_emails(company_name.strip(), _normalise_job_url(job_url))
    return {
        "company_name": company_name,
        "job_url": _normalise_job_url(job_url),
        "candidates": guesses,
    }


@app.get("/resume-profiles", tags=["Campaign"], summary="List available resume profiles")
async def list_resume_profiles() -> dict[str, Any]:
    resume_dir = BASE_DIR / "resume"
    files = []
    if resume_dir.exists():
        files = sorted([p.name for p in resume_dir.iterdir() if p.is_file() and p.suffix.lower() in {".pdf"}])
    return {"profiles": files}


@app.post("/upload-resume", tags=["Campaign"], summary="Upload a resume PDF")
async def upload_resume(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided.")
    suffix = FilePath(file.filename).suffix.lower()
    if suffix != ".pdf":
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")
    configured = FilePath(settings.candidate_resume_path)
    dest = configured if configured.is_absolute() else (BASE_DIR / configured)
    dest.parent.mkdir(parents=True, exist_ok=True)
    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:  # 20 MB hard limit
        raise HTTPException(status_code=413, detail="File too large (max 20 MB).")
    try:
        dest.write_bytes(contents)
    except OSError as exc:
        logger.error(f"[Upload] Could not save resume to {dest}: {exc}")
        raise HTTPException(
            status_code=500,
            detail=f"Could not save resume on server: {exc}",
        )
    logger.info(f"[Upload] Resume saved to {dest} ({len(contents)} bytes)")
    return {"saved_as": dest.name, "size_bytes": len(contents)}


@app.post("/followups/run-due", tags=["Campaign"], summary="Send due follow-up emails")
async def run_due_followups(request: Request) -> dict[str, Any]:
    _require_send_authorization(request)

    now = datetime.now(timezone.utc)
    sent = 0
    failed = 0
    skipped = 0

    for item in FOLLOWUP_QUEUE:
        if item.get("status") != "scheduled":
            continue
        due_at_raw = item.get("due_at")
        try:
            due_at = datetime.fromisoformat(str(due_at_raw))
        except Exception:
            item["status"] = "invalid"
            skipped += 1
            continue

        if due_at > now:
            continue

        if any(
            reply.get("company_id") == item.get("company_id")
            and reply.get("status") in {"replied", "interested"}
            for reply in REPLY_EVENTS
        ):
            item["status"] = "stopped_on_reply"
            skipped += 1
            continue

        result = await send_email(
            to_address=item["recipient_email"],
            subject=item["subject"],
            body=item["body"],
            attach_resume=False,
        )
        item["status"] = "sent" if result.success else "failed"
        item["message"] = result.message
        if result.success:
            sent += 1
        else:
            failed += 1

    return {
        "sent": sent,
        "failed": failed,
        "skipped": skipped,
        "queue_size": len(FOLLOWUP_QUEUE),
    }


@app.post("/replies/sync", tags=["Campaign"], summary="Sync reply events")
async def sync_replies(events: list[dict[str, Any]]) -> dict[str, Any]:
    accepted = 0
    for event in events:
        company_id = str(event.get("company_id", "")).strip()
        status_text = str(event.get("status", "replied")).strip().lower()
        if not company_id:
            continue
        REPLY_EVENTS.append(
            {
                "company_id": company_id,
                "status": status_text,
                "message": str(event.get("message", "")).strip(),
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
        )
        accepted += 1
    return {"accepted": accepted, "total_events": len(REPLY_EVENTS)}


@app.get("/analytics/overview", tags=["Campaign"], summary="Campaign analytics summary")
async def analytics_overview() -> dict[str, Any]:
    companies = db.list_companies()
    logs = db.list_email_logs()
    sent = [x for x in logs if str(x.get("status", "")).lower() == "sent"]
    failed = [x for x in logs if str(x.get("status", "")).lower() == "failed"]
    pending = [x for x in logs if str(x.get("status", "")).lower() == "pending"]

    replies = [r for r in REPLY_EVENTS if r.get("status") in {"replied", "interested", "positive"}]
    positive = [r for r in REPLY_EVENTS if r.get("status") in {"interested", "positive"}]

    sent_count = len(sent)
    reply_rate = (len(replies) / sent_count * 100.0) if sent_count else 0.0
    positive_rate = (len(positive) / sent_count * 100.0) if sent_count else 0.0

    return {
        "totals": {
            "companies": len(companies),
            "sent": sent_count,
            "failed": len(failed),
            "pending": len(pending),
            "replies": len(replies),
            "positive_replies": len(positive),
            "followups_scheduled": len([f for f in FOLLOWUP_QUEUE if f.get("status") == "scheduled"]),
        },
        "rates": {
            "reply_rate_percent": round(reply_rate, 2),
            "positive_reply_rate_percent": round(positive_rate, 2),
        },
    }


@app.get("/health", summary="Health check", tags=["Meta"])
async def health() -> dict[str, str]:
    """Simple liveness probe — returns 200 OK when the service is running."""
    return {"status": "ok"}


@app.get("/dashboard-data", tags=["Meta"], summary="Dashboard aggregate data")
async def dashboard_data() -> dict[str, Any]:
    """Return companies and email logs in one request for the web dashboard."""
    companies: list[dict[str, Any]] = []
    email_logs: list[dict[str, Any]] = []
    warnings: list[str] = []

    try:
        companies = db.list_companies()
    except Exception as exc:
        warnings.append(f"companies query failed: {exc}")

    try:
        email_logs = db.list_email_logs()
    except Exception as exc:
        warnings.append(f"email_logs query failed: {exc}")

    sent_count = len([x for x in email_logs if str(x.get("status", "")).lower() == "sent"])
    reply_count = len([x for x in REPLY_EVENTS if x.get("status") in {"replied", "interested", "positive"}])
    reply_rate = round((reply_count / sent_count * 100.0), 2) if sent_count else 0.0

    return {
        "companies": companies,
        "email_logs": email_logs,
        "templates": sorted(TEMPLATE_LIBRARY.keys()),
        "followups": {
            "total": len(FOLLOWUP_QUEUE),
            "scheduled": len([f for f in FOLLOWUP_QUEUE if f.get("status") == "scheduled"]),
        },
        "replies": {
            "total": len(REPLY_EVENTS),
            "reply_rate_percent": reply_rate,
        },
        "warnings": warnings,
    }


@app.get("/", include_in_schema=False)
async def root_ui() -> FileResponse:
    """Serve React dashboard build if present, else fallback to legacy template."""
    react_index = FRONTEND_DIST_DIR / "index.html"
    if react_index.exists():
        return FileResponse(react_index)
    return FileResponse(BASE_DIR / "templates" / "index.html")


@app.get("/favicon.ico", include_in_schema=False)
async def favicon() -> Response:
    """Return an empty favicon response to avoid browser 404 noise."""
    return Response(status_code=204)


app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
app.mount("/assets", StaticFiles(directory=FRONTEND_ASSETS_DIR, check_dir=False), name="assets")
